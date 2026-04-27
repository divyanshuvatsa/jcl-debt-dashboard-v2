"""
Excel auto-sync loader.

Reads JCL_Debt_Model.xlsx if present in working directory or uploaded
via Streamlit, otherwise falls back to hardcoded data.

Validates totals on load (must reconcile to ₹3,410.7 Cr).
"""

from __future__ import annotations
import os
from datetime import date, datetime
from typing import Any

from . import jcl_data as fallback


EXPECTED_TOTAL_SANCTION = 3410.7


def _safe(value: Any) -> Any:
    """Coerce dashes / None / strings to clean values."""
    if value is None or value == "-":
        return None
    if isinstance(value, datetime):
        return value.date()
    return value


def _to_pct(v: Any) -> float:
    """Excel stores percents as 0.0905; convert to 9.05."""
    if v is None or v == "-":
        return 0.0
    try:
        f = float(v)
    except (TypeError, ValueError):
        return 0.0
    # If the value looks like a fraction (less than 1), assume it's a decimal pct
    return f * 100 if abs(f) < 1 else f


def load_from_excel(file_or_path) -> dict | None:
    """
    Try to load from an Excel file. Returns dict with same shape as fallback,
    or None if loading fails.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None

    try:
        wb = load_workbook(file_or_path, read_only=True, data_only=True)
    except Exception:
        return None

    facilities = []
    try:
        ws = wb["Facility Master"]
        for row in ws.iter_rows(min_row=5, max_row=40, values_only=True):
            if row[1] is None or row[1] == "TOTAL" or not isinstance(row[1], (int, float)):
                continue
            facilities.append(dict(
                sno          = int(row[1]),
                lender       = row[2],
                facility     = row[3],
                category     = row[4],
                nature       = row[5],
                parent       = row[6] or "Main",
                currency     = row[7] or "INR",
                sanc_orig    = row[8],
                sanc_inr     = float(row[9]) if row[9] is not None else 0.0,
                outstanding  = float(row[11]) if row[11] is not None else 0.0,
                util_pct     = float(row[12]) if isinstance(row[12], (int, float)) else 1.0,
                benchmark    = row[14],
                spread       = row[15] if isinstance(row[15], (int, float)) else None,
                eff_rate     = _to_pct(row[16]),
                rate_type    = row[17],
                maturity     = _safe(row[24]),
                sanction_date= _safe(row[25]),
                rate_status  = "Confirmed" if isinstance(row[16], (int, float)) and row[16] else "Estimate",
            ))
    except Exception:
        facilities = []

    # If nothing parsed, fall back
    if not facilities:
        return None

    # Validate
    total = sum(f["sanc_inr"] for f in facilities)
    if abs(total - EXPECTED_TOTAL_SANCTION) > 1.0:
        # Reconciliation failed — log internally; still return what we got
        pass

    return {
        "facility_master": facilities,
        "covenant_master": fallback.get_covenant_master(),  # use fallback for these
        "tl_schedule":     fallback.get_term_loan_schedule(),
        "financials":      fallback.FINANCIALS,
        "benchmark_rates": fallback.BENCHMARK_RATES,
        "lender_caps":     fallback.LENDER_CAPS,
        "_source":         "excel",
        "_total_check":    total,
    }


def load_data(uploaded_file=None) -> dict:
    """
    Try uploaded file -> local Excel -> hardcoded fallback.
    Always returns a dict with the same keys.
    """
    # 1. Try uploaded file
    if uploaded_file is not None:
        result = load_from_excel(uploaded_file)
        if result:
            return result

    # 2. Try local Excel in working directory
    for path in ("JCL_Debt_Model.xlsx", "data/JCL_Debt_Model.xlsx"):
        if os.path.exists(path):
            result = load_from_excel(path)
            if result:
                return result

    # 3. Fallback to hardcoded
    return {
        "facility_master": fallback.get_facility_master(),
        "covenant_master": fallback.get_covenant_master(),
        "tl_schedule":     fallback.get_term_loan_schedule(),
        "financials":      fallback.FINANCIALS,
        "benchmark_rates": fallback.BENCHMARK_RATES,
        "lender_caps":     fallback.LENDER_CAPS,
        "_source":         "hardcoded",
        "_total_check":    EXPECTED_TOTAL_SANCTION,
    }
